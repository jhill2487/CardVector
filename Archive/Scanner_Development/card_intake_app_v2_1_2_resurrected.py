"""
Putnam Collectibles Card Intake App v2.1.2 Resurrected + DB/CSV Improvements

Purpose:
  Resurrect the Version 2.1.2 OCR/matching behavior that worked well on Watchog-era tests,
  while adding modern lookup loading from Excel, CSV, folders of CSVs, and SQLite.

Still preserved from v2.1.2:
  - Raw/default Tesseract OCR is attempted FIRST.
  - Multiple OCR passes are combined instead of trusting one pass.
  - Card number + card name + OCR token evidence scoring.
  - Database-name rescue path.
  - Optional OCR debug text files.
  - Processed/review folder movement.

New in resurrected build:
  - --lookup accepts .xlsx, .xlsm, .csv, or a folder of CSV files.
  - --sqlite accepts putnam_pokemon_cloud_ready.sqlite style databases.
  - Auto-detects common lookup/database files if --lookup/--sqlite are not provided.
  - Flexible column mapping for TradingCardDex, TCGCSV-style, and Putnam schemas.
  - Card-number normalization handles 073/086 vs 73/86, TG/GG/SV prefixes, and plain numbers.
  - Optional scan-only mode when you want plain text results without touching the workbook.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

try:
    from PIL import Image, ImageOps, ImageFilter, ImageEnhance
except Exception:
    Image = None
    ImageOps = None
    ImageFilter = None
    ImageEnhance = None

try:
    import pytesseract
    DEFAULT_TESSERACT = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(DEFAULT_TESSERACT):
        pytesseract.pytesseract.tesseract_cmd = DEFAULT_TESSERACT
except Exception:
    pytesseract = None

try:
    import cv2
    import numpy as np
except Exception:
    cv2 = None
    np = None

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
LOOKUP_SHEET = "Pokemon_Lookup_Database"
INTAKE_SHEET = "Card Intake"
REQUIRED_COLS = ["Set Name", "Card Name", "Card Number", "Rarity"]

COLUMN_ALIASES = {
    "Set Name": [
        "Set Name", "set_name", "setName", "set", "Set", "groupName", "Group Name",
        "expansion", "Expansion", "series", "Series"
    ],
    "Card Name": [
        "Card Name", "card_name", "name", "Name", "productName", "Product Name",
        "cardName", "CardName", "title", "Title"
    ],
    "Card Number": [
        "Card Number", "card_number", "number", "Number", "cardNumber", "CardNumber",
        "collectorNumber", "Collector Number", "extNumber", "Ext Number", "number_printed"
    ],
    "Rarity": [
        "Rarity", "rarity", "rarityName", "Rarity Name", "rarity_name", "subTypeName"
    ],
}

SET_CODE_HINTS = {
    "CRI": "Chaos Rising",
    "PRE": "Prismatic Evolutions",
    "PAL": "Paldea Evolved",
    "SVI": "Scarlet & Violet",
    "OBF": "Obsidian Flames",
    "MEG": "Mega Evolution",
}


@dataclass
class MatchResult:
    set_name: str
    card_name: str
    card_number: str
    rarity: str
    confidence: float
    reason: str


@dataclass
class OCRResult:
    text: str
    method: str
    card_number: str
    hp: str
    likely_name: str
    evidence_terms: List[str]
    set_code_hint: str = ""


def normalize_text(value: str) -> str:
    value = str(value or "").lower()
    value = value.replace("é", "e").replace("’", "'").replace("‘", "'")
    value = re.sub(r"[^a-z0-9/' ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_col_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def safe_cell(row: pd.Series, col: str) -> str:
    if col not in row.index:
        return ""
    value = row[col]
    if pd.isna(value):
        return ""
    return str(value).strip()


def canonical_card_number(value: str) -> str:
    """Normalize collector numbers for matching: 073/086 -> 73/86, TG25/TG30 preserved."""
    raw = str(value or "").upper().strip()
    raw = raw.replace("#", "")
    raw = re.sub(r"\s+", "", raw)
    if not raw:
        return ""

    def clean_part(part: str) -> str:
        m = re.fullmatch(r"([A-Z]*)(\d+)([A-Z]*)", part)
        if not m:
            return part
        prefix, digits, suffix = m.groups()
        return f"{prefix}{int(digits)}{suffix}"

    if "/" in raw:
        left, right = raw.split("/", 1)
        return f"{clean_part(left)}/{clean_part(right)}"
    return clean_part(raw)


def card_number_left(value: str) -> str:
    canon = canonical_card_number(value)
    return canon.split("/", 1)[0] if canon else ""


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map many public CSV/database schemas into Putnam's four lookup columns."""
    if df.empty:
        return pd.DataFrame(columns=REQUIRED_COLS)

    norm_to_actual = {normalize_col_name(c): c for c in df.columns}
    mapped: Dict[str, pd.Series] = {}

    for target, aliases in COLUMN_ALIASES.items():
        actual = None
        for alias in aliases:
            key = normalize_col_name(alias)
            if key in norm_to_actual:
                actual = norm_to_actual[key]
                break
        if actual is not None:
            mapped[target] = df[actual]
        else:
            mapped[target] = pd.Series([""] * len(df))

    out = pd.DataFrame(mapped)
    for col in REQUIRED_COLS:
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out.dropna(subset=["Card Name"], how="all")
    out = out[out["Card Name"].astype(str).str.strip() != ""]
    out["Card Number Key"] = out["Card Number"].map(canonical_card_number)
    out["Card Number Left Key"] = out["Card Number"].map(card_number_left)
    out["Card Name Key"] = out["Card Name"].map(normalize_text)
    out["Set Name Key"] = out["Set Name"].map(normalize_text)
    return out.reset_index(drop=True)


def read_csv_flexible(path: Path) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, dtype=str)


def load_lookup_from_excel(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    sheet = LOOKUP_SHEET if LOOKUP_SHEET in xls.sheet_names else xls.sheet_names[0]
    return map_columns(pd.read_excel(path, sheet_name=sheet, dtype=str))


def load_lookup_from_csv(path: Path) -> pd.DataFrame:
    return map_columns(read_csv_flexible(path))


def load_lookup_from_csv_folder(path: Path) -> pd.DataFrame:
    frames = []
    for csv_path in sorted(path.rglob("*.csv")):
        try:
            df = read_csv_flexible(csv_path)
            mapped = map_columns(df)
            if not mapped.empty:
                # If no set name exists in the CSV, use folder/file stem as a useful hint.
                if mapped["Set Name"].eq("").all():
                    mapped["Set Name"] = csv_path.parent.name if csv_path.parent != path else csv_path.stem
                    mapped["Set Name Key"] = mapped["Set Name"].map(normalize_text)
                frames.append(mapped)
        except Exception as exc:
            print(f"Warning: skipped CSV {csv_path}: {exc}")
    if not frames:
        return pd.DataFrame(columns=REQUIRED_COLS + ["Card Number Key", "Card Number Left Key", "Card Name Key", "Set Name Key"])
    return pd.concat(frames, ignore_index=True).drop_duplicates(
        subset=["Set Name", "Card Name", "Card Number", "Rarity"]
    ).reset_index(drop=True)


def sqlite_tables(con: sqlite3.Connection) -> List[str]:
    rows = con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
    return [r[0] for r in rows]


def load_lookup_from_sqlite(path: Path, table: Optional[str] = None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"SQLite database not found: {path}")
    con = sqlite3.connect(path)
    try:
        tables = sqlite_tables(con)
        preferred = ["pokemon_cards", "cards", "pokemon_lookup", "products"]
        selected = table or next((t for t in preferred if t in tables), None)
        if not selected:
            # choose first table with a likely name column
            for t in tables:
                info = pd.read_sql_query(f'PRAGMA table_info("{t}")', con)
                cols = {normalize_col_name(c) for c in info["name"].tolist()}
                if {"cardname"} & cols or {"name", "productname"} & cols:
                    selected = t
                    break
        if not selected:
            raise ValueError(f"Could not find a card table. Tables found: {tables}")
        df = pd.read_sql_query(f'SELECT * FROM "{selected}"', con)
        mapped = map_columns(df)
        if mapped.empty:
            raise ValueError(f"SQLite table {selected} loaded but no card rows mapped.")
        return mapped
    finally:
        con.close()


def auto_find_lookup(base: Path) -> Tuple[Optional[Path], Optional[Path]]:
    """Return (lookup_path, sqlite_path)."""
    search_roots = [base, base / "database", base.parent, Path.cwd(), Path.cwd() / "database"]
    sqlite_names = ["putnam_pokemon_cloud_ready.sqlite", "pokemon_cards.sqlite", "pokemon_lookup.sqlite"]
    lookup_names = [
        "Pokemon_Lookup_Database.xlsx",
        "Putnam_Pokemon_Lookup_Database_v3.xlsx",
        "pokemon_cards.csv",
        "putnam_pokemon_cards_cloud_ready.csv",
    ]
    found_sqlite = None
    found_lookup = None
    for root in search_roots:
        for name in sqlite_names:
            p = root / name
            if p.exists() and found_sqlite is None:
                found_sqlite = p
        for name in lookup_names:
            p = root / name
            if p.exists() and found_lookup is None:
                found_lookup = p
    return found_lookup, found_sqlite


def load_lookup_source(workbook_path: Path, lookup_path: Optional[Path], sqlite_path: Optional[Path], sqlite_table: Optional[str]) -> pd.DataFrame:
    if sqlite_path:
        lookup = load_lookup_from_sqlite(sqlite_path, sqlite_table)
        print(f"Lookup loaded from SQLite: {sqlite_path} ({len(lookup):,} cards)")
        return lookup
    if lookup_path:
        if lookup_path.is_dir():
            lookup = load_lookup_from_csv_folder(lookup_path)
        elif lookup_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            lookup = load_lookup_from_excel(lookup_path)
        elif lookup_path.suffix.lower() == ".csv":
            lookup = load_lookup_from_csv(lookup_path)
        else:
            raise ValueError(f"Unsupported lookup source: {lookup_path}")
        print(f"Lookup loaded from {lookup_path}: {len(lookup):,} cards")
        return lookup

    auto_lookup, auto_sqlite = auto_find_lookup(workbook_path.parent if workbook_path else Path.cwd())
    if auto_sqlite:
        lookup = load_lookup_from_sqlite(auto_sqlite, sqlite_table)
        print(f"Lookup auto-loaded from SQLite: {auto_sqlite} ({len(lookup):,} cards)")
        return lookup
    if auto_lookup:
        return load_lookup_source(workbook_path, auto_lookup, None, sqlite_table)

    # Legacy fallback: v2.1.2 lookup sheet inside the workbook.
    if workbook_path and workbook_path.exists():
        lookup = load_lookup_from_excel(workbook_path)
        print(f"Lookup loaded from workbook sheet: {workbook_path} ({len(lookup):,} cards)")
        return lookup

    raise FileNotFoundError(
        "No lookup source found. Provide --lookup CSV/XLSX/folder or --sqlite database\\putnam_pokemon_cloud_ready.sqlite"
    )


def extract_card_number(text: str) -> str:
    patterns = [
        r"\bTG\s*\d{1,3}\s*/\s*TG\s*\d{1,3}\b",
        r"\bGG\s*\d{1,3}\s*/\s*GG\s*\d{1,3}\b",
        r"\bSV\s*\d{1,3}\s*/\s*SV\s*\d{1,3}\b",
        r"\b[A-Z]{1,4}\s*\d{1,3}\s*/\s*[A-Z]{1,4}\s*\d{1,3}\b",
        r"\b\d{1,3}\s*/\s*\d{1,3}\b",
        r"\b(?:No\.?|#)\s*\d{1,3}\b",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            value = re.sub(r"\s+", "", m.group(0).upper())
            value = re.sub(r"^(NO\.?|#)", "", value)
            return canonical_card_number(value)
    return ""


def extract_set_code_hint(text: str) -> str:
    upper = str(text or "").upper()
    # modern English cards often show small set code text such as PAL EN, PRE EN, CRI EN.
    for code in SET_CODE_HINTS:
        if re.search(rf"\b{re.escape(code)}\s*(?:EN)?\b", upper):
            return code
    return ""


def extract_hp(text: str) -> str:
    m = re.search(r"\b(?:HP\s*)?(\d{2,3})\s*HP\b|\bHP\s*(\d{2,3})\b", text, flags=re.IGNORECASE)
    if not m:
        return ""
    return next((g for g in m.groups() if g), "")


def likely_card_name_from_ocr(text: str) -> str:
    bad_words = {
        "basic", "stage", "evolves", "pokemon", "pokémon", "trainer", "energy",
        "weakness", "resistance", "retreat", "illustration", "illus", "copyright",
        "regulation", "rule", "rules", "evolvesfrom", "ability"
    }
    for raw_line in text.splitlines():
        line = raw_line.strip()
        norm = normalize_text(line)
        if not norm:
            continue
        if len(norm) < 3 or len(norm) > 44:
            continue
        if any(w in norm.split() for w in bad_words):
            continue
        if re.fullmatch(r"[0-9 /#.-]+", norm):
            continue
        if len(norm.split()) <= 5:
            return line
    return ""


def extract_evidence_terms(text: str) -> List[str]:
    terms: List[str] = []
    ignored = {
        "the", "and", "your", "you", "this", "that", "card", "cards", "opponent",
        "pokemon", "pokémon", "damage", "during", "turn", "attach", "energy",
        "weakness", "resistance", "retreat", "heads", "tails", "shuffle",
    }
    for raw_line in text.splitlines():
        norm = normalize_text(raw_line)
        if not norm:
            continue
        words = norm.split()
        if 1 <= len(words) <= 4 and not all(w in ignored for w in words):
            if not re.search(r"\d{2,3}\s*hp|\d+\s*/\s*\d+", norm, flags=re.I):
                terms.append(norm)
    out: List[str] = []
    for t in terms:
        if t not in out:
            out.append(t)
    return out[:12]


def preprocess_images(image_path: Path) -> List[Tuple[str, object]]:
    variants: List[Tuple[str, object]] = []
    if Image is None:
        return variants
    img = Image.open(image_path).convert("RGB")
    variants.append(("original", img))

    gray = ImageOps.grayscale(img)
    variants.append(("grayscale", gray))
    enhanced = ImageEnhance.Contrast(gray).enhance(2.0).filter(ImageFilter.SHARPEN)
    variants.append(("pil_contrast_sharpen_2x", enhanced.resize((enhanced.width * 2, enhanced.height * 2))))

    # gentle crop passes keep v2 behavior but help card top/bottom text without requiring full contour logic.
    w, h = img.size
    if w > 0 and h > 0:
        top = img.crop((0, 0, w, int(h * 0.28)))
        bottom = img.crop((0, int(h * 0.72), w, h))
        variants.append(("top_band_raw", top.resize((top.width * 2, top.height * 2))))
        variants.append(("bottom_band_raw", bottom.resize((bottom.width * 2, bottom.height * 2))))

    if cv2 is not None and np is not None:
        arr = np.array(gray)
        arr = cv2.resize(arr, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        arr = cv2.GaussianBlur(arr, (3, 3), 0)
        thresh = cv2.adaptiveThreshold(arr, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 11)
        variants.append(("opencv_adaptive_threshold_2x", Image.fromarray(thresh)))
    return variants


def ocr_image(image_path: Path, save_ocr_text: bool = False, ocr_dir: Optional[Path] = None) -> OCRResult:
    if pytesseract is None or Image is None:
        return OCRResult("", "unavailable", "", "", "", [])

    attempts: List[Tuple[str, str, int]] = []
    configs = ["--psm 6", "--psm 11", "--psm 12", "--psm 7"]

    # Critical v2.1.2 behavior: raw/default first, v1-style.
    try:
        raw_img = Image.open(image_path)
        raw_text = pytesseract.image_to_string(raw_img)
        raw_norm = normalize_text(raw_text)
        if raw_norm:
            raw_score = 1000 + min(len(raw_norm), 300)
            if extract_card_number(raw_text):
                raw_score += 500
            if extract_hp(raw_text):
                raw_score += 50
            attempts.append(("raw_default_v1_style", raw_text, raw_score))
    except Exception as exc:
        attempts.append(("raw_default_v1_style_error", str(exc), 1))

    try:
        for method, img in preprocess_images(image_path):
            for config in configs:
                label = f"{method} {config}"
                try:
                    text = pytesseract.image_to_string(img, config=config)
                except Exception:
                    continue
                norm = normalize_text(text)
                if not norm:
                    continue
                score = min(len(norm), 300)
                if extract_card_number(text):
                    score += 500
                if extract_hp(text):
                    score += 50
                if likely_card_name_from_ocr(text):
                    score += 100
                if "top_band" in method or "bottom_band" in method:
                    score += 20
                attempts.append((label, text, score))
    except Exception:
        pass

    if not attempts:
        return OCRResult("", "none", "", "", "", [])

    attempts.sort(key=lambda x: x[2], reverse=True)
    best_method, best_text, _ = attempts[0]

    combined_parts: List[str] = []
    seen_norms = set()
    for label, text, _score in attempts:
        norm = normalize_text(text)
        if norm and norm not in seen_norms:
            seen_norms.add(norm)
            combined_parts.append(f"[{label}]\n{text}")
    combined_text = "\n\n".join(combined_parts)

    card_number = extract_card_number(combined_text)
    hp = extract_hp(combined_text)
    likely_name = likely_card_name_from_ocr(best_text) or likely_card_name_from_ocr(combined_text)
    evidence_terms = extract_evidence_terms(combined_text)
    set_code_hint = extract_set_code_hint(combined_text)

    if save_ocr_text:
        target_dir = ocr_dir or Path("ocr_debug")
        target_dir.mkdir(parents=True, exist_ok=True)
        out = target_dir / f"{image_path.stem}.ocr.txt"
        out.write_text(
            f"BEST_METHOD: {best_method}\n"
            f"CARD_NUMBER: {card_number}\n"
            f"HP: {hp}\n"
            f"LIKELY_NAME: {likely_name}\n"
            f"SET_CODE_HINT: {set_code_hint}\n\n"
            f"COMBINED_OCR_TEXT:\n{combined_text}",
            encoding="utf-8",
        )

    return OCRResult(combined_text, f"combined OCR; best={best_method}", card_number, hp, likely_name, evidence_terms, set_code_hint)


def score_candidate(ocr: OCRResult, row: pd.Series, full_text_norm: str, filename_norm: str) -> Tuple[float, List[str]]:
    reasons: List[str] = []
    card_name = safe_cell(row, "Card Name")
    set_name = safe_cell(row, "Set Name")
    card_number = safe_cell(row, "Card Number")
    name_norm = safe_cell(row, "Card Name Key") or normalize_text(card_name)
    set_norm = safe_cell(row, "Set Name Key") or normalize_text(set_name)
    lookup_num = safe_cell(row, "Card Number Key") or canonical_card_number(card_number)
    lookup_left = safe_cell(row, "Card Number Left Key") or card_number_left(card_number)
    detected_num = canonical_card_number(ocr.card_number)
    detected_left = card_number_left(ocr.card_number)

    score = 0.0

    if detected_num and lookup_num and detected_num == lookup_num:
        score += 0.68
        reasons.append(f"card number {ocr.card_number}")
    elif detected_left and lookup_left and detected_left == lookup_left:
        score += 0.45
        reasons.append(f"card number left side {detected_left}")

    if name_norm and name_norm in full_text_norm:
        score += 0.45
        reasons.append("card name in OCR")

    likely_name_norm = normalize_text(ocr.likely_name)
    if likely_name_norm and name_norm:
        sim = SequenceMatcher(None, likely_name_norm, name_norm).ratio()
        if sim >= 0.82:
            score += 0.35 * sim
            reasons.append(f"top-line name similarity {sim:.2f}")
        elif sim >= 0.65:
            score += 0.18 * sim
            reasons.append(f"possible top-line name {sim:.2f}")

    if name_norm:
        name_tokens = [t for t in name_norm.split() if len(t) >= 3]
        if name_tokens:
            token_hits = sum(1 for t in name_tokens if t in full_text_norm)
            if token_hits:
                score += 0.12 * (token_hits / len(name_tokens))
                reasons.append(f"name token hits {token_hits}/{len(name_tokens)}")
        whole_name_score = SequenceMatcher(None, full_text_norm[:800], name_norm).ratio()
        score += min(0.12, whole_name_score * 0.15)

    if filename_norm and name_norm and name_norm in filename_norm:
        score += 0.10
        reasons.append("filename hint")

    if set_norm and (set_norm in full_text_norm or set_norm in filename_norm):
        score += 0.12
        reasons.append("set name hint")

    if ocr.set_code_hint and SET_CODE_HINTS.get(ocr.set_code_hint):
        hinted_set = normalize_text(SET_CODE_HINTS[ocr.set_code_hint])
        if hinted_set and hinted_set == set_norm:
            score += 0.18
            reasons.append(f"set code {ocr.set_code_hint}")

    return min(score, 0.99), reasons


def match_card(ocr: OCRResult, lookup: pd.DataFrame, filename_hint: str = "") -> Optional[MatchResult]:
    full_text_norm = normalize_text(ocr.text + " " + filename_hint)
    filename_norm = normalize_text(filename_hint)
    if not full_text_norm:
        return None

    candidates = lookup.copy()

    # set-code narrowing when card number was also detected; avoids duplicate number collisions.
    if ocr.set_code_hint and SET_CODE_HINTS.get(ocr.set_code_hint):
        hinted_set_key = normalize_text(SET_CODE_HINTS[ocr.set_code_hint])
        hinted_rows = candidates[candidates["Set Name Key"] == hinted_set_key]
        if not hinted_rows.empty:
            candidates = hinted_rows

    exact_name_rows = []
    for _, row in candidates.iterrows():
        name_norm = safe_cell(row, "Card Name Key") or normalize_text(safe_cell(row, "Card Name"))
        if name_norm and len(name_norm) >= 3 and name_norm in full_text_norm:
            exact_name_rows.append(row)
    if len(exact_name_rows) == 1 and not ocr.card_number:
        row = exact_name_rows[0]
        return MatchResult(safe_cell(row, "Set Name"), safe_cell(row, "Card Name"), safe_cell(row, "Card Number"), safe_cell(row, "Rarity"), 0.72, "Exact database card name found in combined OCR text")

    detected_num = canonical_card_number(ocr.card_number)
    detected_left = card_number_left(ocr.card_number)
    candidates_to_score = candidates
    if detected_num:
        same_num = candidates[candidates["Card Number Key"] == detected_num]
        if same_num.empty and detected_left:
            same_num = candidates[candidates["Card Number Left Key"] == detected_left]
        if not same_num.empty:
            candidates_to_score = same_num

    best_row = None
    best_score = 0.0
    second_best_score = 0.0
    best_reasons: List[str] = []

    for _, row in candidates_to_score.iterrows():
        score, reasons = score_candidate(ocr, row, full_text_norm, filename_norm)
        if score > best_score:
            second_best_score = best_score
            best_score = score
            best_row = row
            best_reasons = reasons
        elif score > second_best_score:
            second_best_score = score

    if best_score < 0.50 and len(candidates_to_score) != len(lookup):
        for _, row in lookup.iterrows():
            score, reasons = score_candidate(ocr, row, full_text_norm, filename_norm)
            if score > best_score:
                second_best_score = best_score
                best_score = score
                best_row = row
                best_reasons = reasons
            elif score > second_best_score:
                second_best_score = score

    if best_row is None:
        return None

    ambiguity_gap = best_score - second_best_score
    if not detected_num and ambiguity_gap < 0.08 and best_score < 0.82:
        confidence = max(0.50, best_score - 0.08)
        reason = "Ambiguous name match; " + ", ".join(best_reasons)
    else:
        confidence = best_score
        reason = ", ".join(best_reasons) if best_reasons else "OCR fuzzy match"

    if confidence >= 0.55:
        return MatchResult(
            set_name=safe_cell(best_row, "Set Name"),
            card_name=safe_cell(best_row, "Card Name"),
            card_number=safe_cell(best_row, "Card Number"),
            rarity=safe_cell(best_row, "Rarity"),
            confidence=min(confidence, 0.99),
            reason=reason,
        )
    return None


def next_empty_row(ws) -> int:
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=6).value in (None, ""):
            return row
    return ws.max_row + 1


def append_to_card_intake(workbook_path: Path, image_path: Path, result: Optional[MatchResult], ocr: OCRResult, condition: str, finish: str, qty: int = 1) -> None:
    if not workbook_path.exists():
        print(f"  Intake skipped: workbook not found: {workbook_path}")
        return
    wb = load_workbook(workbook_path)
    if INTAKE_SHEET not in wb.sheetnames:
        print(f"  Intake skipped: workbook does not contain sheet: {INTAKE_SHEET}")
        return
    ws = wb[INTAKE_SHEET]
    row = next_empty_row(ws)
    intake_id = f"INTAKE-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{row}"
    review_status = "Needs Review"
    notes = "No confident match. Review manually."

    if result:
        if result.confidence >= 0.85:
            review_status = "Auto Match - Review"
        elif result.confidence >= 0.70:
            review_status = "Likely Match - Review"
        else:
            review_status = "Low Confidence Match - Review"
        notes = (
            f"V2.1.2 resurrected confidence={result.confidence:.2f}; reason={result.reason}; "
            f"OCR method={ocr.method}; likely OCR name={ocr.likely_name}; "
            f"OCR number={ocr.card_number}; OCR HP={ocr.hp}; OCR set code={ocr.set_code_hint}"
        )

    values: Dict[int, object] = {
        1: intake_id, 2: datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 3: str(image_path),
        4: "Pokemon", 5: result.set_name if result else "", 6: result.card_name if result else "",
        7: result.card_number if result else ocr.card_number, 8: result.rarity if result else "",
        9: condition, 10: finish, 11: qty, 12: "", 13: "", 14: "", 15: "",
        19: review_status, 20: "No", 21: notes,
    }
    for col, value in values.items():
        ws.cell(row=row, column=col).value = value
    wb.save(workbook_path)


def safe_move(src: Path, dst_dir: Path) -> Path:
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / src.name
    if not dst.exists():
        shutil.move(str(src), str(dst))
        return dst
    stem, suffix = src.stem, src.suffix
    i = 1
    while True:
        candidate = dst_dir / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            shutil.move(str(src), str(candidate))
            return candidate
        i += 1


def iter_images(input_path: Path) -> List[Path]:
    if input_path.is_file() and input_path.suffix.lower() in IMAGE_EXTENSIONS:
        return [input_path]
    input_path.mkdir(parents=True, exist_ok=True)
    return sorted([p for p in input_path.iterdir() if p.suffix.lower() in IMAGE_EXTENSIONS])


def process_images(workbook_path: Path, input_path: Path, processed_dir: Path, review_dir: Path, lookup: pd.DataFrame, condition: str, finish: str, save_ocr_text: bool, scan_only: bool, no_move: bool) -> None:
    processed_dir.mkdir(parents=True, exist_ok=True)
    review_dir.mkdir(parents=True, exist_ok=True)
    images = iter_images(input_path)
    if not images:
        print(f"No images found in {input_path}")
        return

    print(f"Images found: {len(images)}")
    for img in images:
        print(f"\nProcessing {img.name}...")
        ocr = ocr_image(img, save_ocr_text=save_ocr_text)
        if ocr.text:
            print(f"  OCR text found using {ocr.method}.")
            if ocr.likely_name:
                print(f"  Likely OCR name: {ocr.likely_name}")
            if ocr.card_number:
                print(f"  OCR card number: {ocr.card_number}")
            if ocr.hp:
                print(f"  OCR HP: {ocr.hp}")
            if ocr.set_code_hint:
                print(f"  OCR set code hint: {ocr.set_code_hint} -> {SET_CODE_HINTS.get(ocr.set_code_hint, '')}")
        else:
            print("  OCR unavailable or no text found.")

        result = match_card(ocr, lookup, img.stem)

        if result:
            print(f"  Match: {result.card_name} | {result.set_name} | {result.card_number} | {result.confidence:.2f}")
            print(f"  Reason: {result.reason}")
        else:
            print("  Match: needs manual review")

        if not scan_only:
            append_to_card_intake(workbook_path, img, result, ocr, condition=condition, finish=finish)

        if not no_move and input_path.is_dir():
            destination = processed_dir if result else review_dir
            moved_to = safe_move(img, destination)
            print(f"  Moved to: {moved_to}")
        else:
            print("  Image left in place.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Putnam scanner v2.1.2 resurrected with SQLite/CSV lookup support.")
    parser.add_argument("--workbook", default="Putnam_Master_Inventory_Card_Intake_Workflow.xlsx", help="Path to master workbook for Card Intake output.")
    parser.add_argument("--lookup", default="", help="Lookup source: .xlsx, .csv, or folder of CSV files.")
    parser.add_argument("--sqlite", default="", help="SQLite lookup database, e.g. database\\putnam_pokemon_cloud_ready.sqlite.")
    parser.add_argument("--sqlite-table", default="", help="Optional SQLite table name.")
    parser.add_argument("--input", default="input_photos", help="Image file or folder containing card photos.")
    parser.add_argument("--processed", default="processed_photos", help="Folder for matched photos.")
    parser.add_argument("--review", default="review_photos", help="Folder for unmatched photos.")
    parser.add_argument("--condition", default="Near Mint", help="Default condition.")
    parser.add_argument("--finish", default="Normal", help="Default finish.")
    parser.add_argument("--save-ocr-text", action="store_true", help="Save OCR debug text files to ocr_debug.")
    parser.add_argument("--scan-only", action="store_true", help="Print results only; do not write Card Intake workbook.")
    parser.add_argument("--no-move", action="store_true", help="Leave images in place after scanning.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    lookup_path = Path(args.lookup) if args.lookup else None
    sqlite_path = Path(args.sqlite) if args.sqlite else None
    lookup = load_lookup_source(workbook_path, lookup_path, sqlite_path, args.sqlite_table or None)

    process_images(
        workbook_path=workbook_path,
        input_path=Path(args.input),
        processed_dir=Path(args.processed),
        review_dir=Path(args.review),
        lookup=lookup,
        condition=args.condition,
        finish=args.finish,
        save_ocr_text=args.save_ocr_text,
        scan_only=args.scan_only,
        no_move=args.no_move,
    )


if __name__ == "__main__":
    main()
